import os

from shared import excel_helpers
import openpyxl as xl
from openpyxl.worksheet import table as xltable
from openpyxl import chart as xlchart
from openpyxl import utils as xlutils


EXCEL_FILENAME = 'income_annual_apha.xlsx'

class IncomeStatementWrapper:
    def __init__(self, income_statement_json, ticker):
        self.income_statement = income_statement_json
        self.ticker = ticker
        self.workbook = xl.Workbook()
        self.worksheet = self.workbook.create_sheet('Income', 0)
        self.worksheet.sheet_properties.tabColor = 'ff9191'

    def open_excel_sheet(self):
        os.startfile(EXCEL_FILENAME)

    # is_quarterly false implies that this is annual
    def generate_excel_file(self):
        annual_reports = self.income_statement['annualReports']
        quarterly_reports = self.income_statement['quarterlyReports']
        annual_column_heading, annual_table_data = self._get_table_data(annual_reports)
        quarterly_column_heading, quarterly_table_data = self._get_table_data(quarterly_reports)

        self.worksheet.append(annual_column_heading)
        for row in annual_table_data:
            self.worksheet.append(row)
        self.worksheet.append(quarterly_column_heading)
        for row in quarterly_table_data:
            self.worksheet.append(row)

        length_of_annual_data = len(annual_table_data[0])
        # Moves the auarterly sheet next to the annual sheet
        self.worksheet.move_range(
            'A6:AAA10',  # Value after : is purposefully very large since we don't care about copying blank values
            rows=-5,  # 5 accounts for column header and 4 data items we are pulling for
            cols=length_of_annual_data + 1,  # add 1 to leave space in between tables
        )
        self.worksheet.add_table(xltable.Table(
            displayName='Annual',
            # If this isn't specified the program will crash in a cryptic way
            ref='A1:{}5'.format(excel_helpers.convert_number_to_cell_value(length_of_annual_data)),
            tableStyleInfo=xltable.TableStyleInfo(
                name='TableStyleMedium9',
                showFirstColumn=True,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        ))
        self.worksheet.add_table(xltable.Table(
            displayName='Quarterly',
            # If this isn't specified the program will crash in a cryptic way
            ref='{}1:AAA5'.format(excel_helpers.convert_number_to_cell_value(length_of_annual_data + 2)),
            tableStyleInfo=xltable.TableStyleInfo(
                name='TableStyleMedium9',
                showFirstColumn=True,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        ))

        annual_chart = self._get_chart(length_of_annual_data)
        length_of_quarterly_data = len(quarterly_table_data[0])
        quarterly_chart = self._get_chart(length_of_quarterly_data, length_of_annual_data + 1)
        # quarterly_chart = self._get_chart(quarterly_table_data, length_of_annual_data)
        self.worksheet.add_chart(annual_chart, 'A10')
        self.worksheet.add_chart(quarterly_chart, 'G10')

        # 100 is a dumb value to just set all the column widths to be LORGE
        for i in range(100):
            self.worksheet.column_dimensions[xlutils.get_column_letter(i + 1)].width = 15

        self.workbook.save(EXCEL_FILENAME)

    def _get_table_data(self, reports):
        # top left most cell shows ticker + currency
        column_headings = ['{} ({})'.format(self.ticker, reports[0]['reportedCurrency'])]
        data = [
            ['Revenue'],
            ['Gross'],
            ['Op Inc'],
            ['Net'],
        ]

        for report in reports:
            fiscal_date_ending = str(report['fiscalDateEnding'])
            column_headings.append(fiscal_date_ending)

            total_revenue = excel_helpers.string_to_float(report['totalRevenue'])
            gross_profit = excel_helpers.string_to_float(report['grossProfit'])
            operating_income = excel_helpers.string_to_float(report['operatingIncome'])
            net_income = excel_helpers.string_to_float(report['netIncome'])

            data[0].append(total_revenue)
            data[1].append(gross_profit)
            data[2].append(operating_income)
            data[3].append(net_income)

        return column_headings, data

    def _get_chart(self, chart_length, offset=0):
        chart = xlchart.AreaChart()
        chart.title='Pretty graph'
        chart.style = 42
        chart.x_axis.scaling.orientation = "maxMin"
        chart.x_axis.title = 'Time'
        chart.y_axis.title = '$'

        yaxis_dates = xlchart.Reference(
            self.worksheet,
            min_col=offset + 1,
            max_col=offset + chart_length + 1,
            min_row=1,
            max_row=1,
        )
        xaxis_data = xlchart.Reference(
            self.worksheet,
            min_col=offset + 1,
            max_col=offset + chart_length + 1,
            min_row=2,
            max_row=5,
        )
        chart.add_data(xaxis_data, titles_from_data=True, from_rows=True)
        chart.set_categories(yaxis_dates)
        return chart

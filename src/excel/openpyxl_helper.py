from openpyxl import chart, utils
from openpyxl.chart import layout
from openpyxl.styles import numbers
from openpyxl.worksheet import table


TABLE_STYLES = {
    'black': 'TableStyleMedium1',
    'blue': 'TableStyleMedium9',
    'purp': 'TableStyleMedium12',
    'red': 'TableStyleMedium3',
    'teal': 'TableStyleMedium13',
    'green': 'TableStyleMedium11',
}

def get_column_pair(row, column):
    return '{}{}'.format(
        utils.get_column_letter(column),
        row,
    )

def add_cell(
    worksheet,
    cell,
    value,
):
    worksheet[cell].value = value

def add_table(
    worksheet,
    name,
    num_rows,
    num_columns,
    style,
    row_offset=1,
    column_offset=1,
    showRowStripes=True,
):
    # Update all cells that contain numbers to be comma separated
    for i in range(row_offset + 1, num_rows + row_offset):
        for j in range(column_offset + 1, num_columns + column_offset):
            cell = worksheet[get_column_pair(i, j)]
            if isinstance(cell.value, float):
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    worksheet.add_table(table.Table(
        displayName=name,
        ref='{}:{}'.format(
            get_column_pair(row_offset, column_offset),
            get_column_pair(num_rows + row_offset - 1, num_columns + column_offset - 1),
        ),
        tableStyleInfo=table.TableStyleInfo(
            name=TABLE_STYLES[style],
            showFirstColumn=True,
            showLastColumn=False,
            showRowStripes=showRowStripes,
            showColumnStripes=False,
        ),
    ))

def add_sheet(
    workbook,
    name,
    tab_color,
    index,
):
    worksheet = workbook.create_sheet(name, index)
    worksheet.sheet_properties.tabColor = tab_color
    # 100 is a dumb value to just set all the column widths to be LORGE
    worksheet.column_dimensions[utils.get_column_letter(1)].width = 20
    for i in range(1, 100):
        worksheet.column_dimensions[utils.get_column_letter(i + 1)].width = 15

    return worksheet

def add_graph(
    worksheet,
    title,
    num_rows,
    num_columns,
    chart_type,
    x_axis_label='',
    y_axis_label='',
    row_offset=1,
):
    area_chart = None
    if chart_type == 'area':
        area_chart = chart.AreaChart()
        area_chart.style = 42
    elif chart_type == 'bar':
        area_chart = chart.BarChart()
        area_chart.style = 42
        area_chart.shape = 4
    elif chart_type == 'line':
        area_chart = chart.LineChart()
        area_chart.style = 42
    else:
        raise Exception('Unknown chart_type param: {}'.format(chart_type))

    area_chart.title=title
    area_chart.x_axis.scaling.orientation = "maxMin"
    area_chart.x_axis.title = x_axis_label
    # Makes the y axis labels vertical instead of slanted so they all fit
    area_chart.y_axis.title = y_axis_label
    area_chart.x_axis.tickLblSkip = 1
    # put legend on the bottom which allows the graph to be wider
    area_chart.legend.position = 'b'
    # not sure exactly how this works but this makes the graph bigger within the space alotted
    area_chart.layout = layout.Layout(
        layout.ManualLayout(
            x=0, y=0,
            h=.8, w=.8,
        ),
    )

    yaxis_dates = chart.Reference(
        worksheet,
        min_col=2,
        max_col=num_columns + 1,
        min_row=row_offset,
        max_row=row_offset,
    )
    xaxis_data = chart.Reference(
        worksheet,
        min_col=1,
        max_col=num_columns + 1,
        min_row=row_offset + 1,
        max_row=row_offset + num_rows - 1,
    )
    area_chart.add_data(xaxis_data, titles_from_data=True, from_rows=True)
    area_chart.set_categories(yaxis_dates)

    # Override default width which is 5 pts, set in EMU units (12700 EMU = 1 pt)
    if chart_type == 'line':
        for line in area_chart.series:
            line.graphicalProperties.line.width = 25400

    start_cell = 'A{}'.format(row_offset + num_rows)
    worksheet.add_chart(area_chart, start_cell)

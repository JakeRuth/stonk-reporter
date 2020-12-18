import collections
import json
import os

import openpyxl as xl
from openpyxl.worksheet import table as xltable
from openpyxl import chart as xlchart
from openpyxl import utils as xlutils

from shared import excel_helpers
from shared import stock_api

# define data shape
stock_data = {
    'Revenue': [],
    'Cost of Revenue': [],
    'Gross Profit': [],
    'Gross Margin': [],
    'Operating Expense': [],
    'Operating Income': [],
    'Operating Margin': [],
    'Revenue Growth y/y': [],
    'Total Cash': [],  # can't find this
    'Current Assets': [],
    'Current Liabilities Total': [],
    'Current Ratio': [],
    'Total Assets': [],
    'Total Liabilities': [],
    'Stock Holder Equity': [],
    'Stock Holder Growth q/q': [],
    'Goodwill': [],
    'Intangible Assets': [],
    'Tangible Assets': [],
    'Total Liabilities': [],
    'Tangible Book Value': [],
    'Free Cash Flow': [],
    'Free Cash Flow (TTM)': [],
    'Valuation': [],
    'P/FCF Ratio': [],
    'P/E Ratio': [],
    'P/S Ratio': [],
}

class TTM_Data(object):
    def __init__(self, stock_data_key, api_field_name):
        self.stock_data_key = stock_data_key
        self.api_field_name = api_field_name
        self.value = 0

# Get input from user
# api_key = input('Enter your ALPHA Vantage API key: ')
# stock_tickers = input('Enter stock tickers (comma separated, no spaces): ')
api_key = 'jkjkj'
stock_tickers = ['APHA']

# Uncomment for local dev
income_quarterly_reports = None
income_annual_reports = None
with open('static_json_data/apha_income.json') as json_file:
    income_statement = json.load(json_file)
    income_quarterly_reports = income_statement['quarterlyReports']
    income_annual_reports = income_statement['annualReports']

balance_quarterly_reports = None
balance_annual_reports = None
with open('static_json_data/apha_balance.json') as json_file:
    balance_statement = json.load(json_file)
    balance_quarterly_reports = balance_statement['quarterlyReports']
    balance_annual_reports = balance_statement['annualReports']

stock_api_wrapper = stock_api.StockApiWrapper(api_key)
heading_row = ['']  # intentially leave first cell blank
# TODO: Actually implement calling all tickers in the stock api and get mapping of data
for ticker in stock_tickers:
    heading_row.append(ticker)
    # income_quarterly_reports = stock_api_wrapper.get_income_statement_json(ticker)['quarterlyReports']
    # cashflow = stock_api_wrapper.get_cashflow_json(ticker)
    # company_overview = stock_api_wrapper.get_company_overview_json(ticker)

    ttm_income_data_list = {
        'Revenue': TTM_Data('Revenue', 'totalRevenue'),
        'Cost of Revenue': TTM_Data('Cost of Revenue', 'costOfRevenue'),
        'Gross Profit': TTM_Data('Gross Profit', 'grossProfit'),
        'Operating Income': TTM_Data('Operating Income', 'operatingIncome'),
    }
    for qr in income_quarterly_reports[:4]:
        for ttm_data in ttm_income_data_list.values():
            ttm_data.value = ttm_data.value + excel_helpers.string_to_float(qr[ttm_data.api_field_name])

    for ttm_data in ttm_income_data_list.values():
        new_value = ttm_data.value / 1000
        stock_data[ttm_data.stock_data_key].append(new_value)

    # for some reason the api doesn't return this for us
    operating_expense = ttm_income_data_list['Gross Profit'].value - ttm_income_data_list['Operating Income'].value
    stock_data['Operating Expense'].append(operating_expense / 1000)

    gross_margin = ttm_income_data_list['Gross Profit'].value / ttm_income_data_list['Revenue'].value
    stock_data['Gross Margin'].append(round(gross_margin, 3))

    operating_margin = ttm_income_data_list['Operating Income'].value / ttm_income_data_list['Revenue'].value
    stock_data['Operating Margin'].append(round(operating_margin, 3))

    revenue_year_one = excel_helpers.string_to_float(income_annual_reports[0]['totalRevenue'])
    revenue_year_two = excel_helpers.string_to_float(income_annual_reports[1]['totalRevenue'])
    renevue_growth_yoy = (revenue_year_one - revenue_year_two) / revenue_year_one
    stock_data['Revenue Growth y/y'].append(round(renevue_growth_yoy, 3))

    # TODO: Get total cash from "cash" in the balance sheet
    last_report = balance_quarterly_reports[0]
    stock_data['Total Cash'].append(excel_helpers.string_to_float(last_report['cash']) / 1000)
    stock_data['Current Assets'].append(excel_helpers.string_to_float(last_report['totalCurrentAssets']) / 1000)
    stock_data['Current Liabilities Total'].append(excel_helpers.string_to_float(last_report['totalCurrentLiabilities']) / 1000)
    stock_data['Total Assets'].append(excel_helpers.string_to_float(last_report['totalAssets']) / 1000)
    stock_data['Total Liabilities'].append(excel_helpers.string_to_float(last_report['totalLiabilities']) / 1000)
    stock_data['Goodwill'].append(excel_helpers.string_to_float(last_report['goodwill']) / 1000)
    stock_data['Intangible Assets'].append(excel_helpers.string_to_float(last_report['intangibleAssets']) / 1000)
    stock_data['Tangible Assets'].append(excel_helpers.string_to_float(last_report['netTangibleAssets']) / 1000)

workbook = xl.Workbook()
worksheet = workbook.create_sheet('Overview', 0)
worksheet.append(heading_row)
for key in stock_data.keys():
    row_data = [key] + stock_data[key]
    worksheet.append(row_data)

worksheet.add_table(xltable.Table(
    displayName='Overview',
    # If this isn't specified the program will crash in a cryptic way
    ref='A1:{}{}'.format(
        xlutils.get_column_letter(len(stock_tickers) + 1),
        len(stock_data.keys()) + 1,
    ),
    tableStyleInfo=xltable.TableStyleInfo(
        name='TableStyleMedium3',
        showFirstColumn=True,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
))

# 100 is a dumb value to just set all the column widths to be LORGE
for i in range(100):
    worksheet.column_dimensions[xlutils.get_column_letter(i + 1)].width = 25

workbook.save('pronkSheet.xlsx')
os.startfile('pronkSheet.xlsx')

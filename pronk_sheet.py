import json
import os
import time

import openpyxl as xl
from openpyxl.worksheet import table as xltable
from openpyxl import chart as xlchart
from openpyxl import utils as xlutils

from shared import excel_helpers
from shared import stock_api

stock_data = {
    'Revenue TTM': [],
    'Cost of Revenue TTM': [],
    'Gross Profit TTM': [],
    'Gross Margin TTM': [],
    'Operating Expense TTM': [],
    'Operating Income TTM': [],
    'Operating Margin TTM': [],
    'Revenue Growth y/y': [],
    'Total Cash': [],
    'Current Assets': [],
    'Current Liabilities Total': [],
    'Current Ratio': [],
    'Total Assets': [],
    'Total Liabilities': [],
    'Stock Holder Equity': [],
    'Stock Holder Growth q/q': [],
    'Goodwill': [],
    'Intangible Assets': [],
    'Tangible Assets': [],
    'Total Liabilities': [],
    'Tangible Book Value': [],
    'Free Cash Flow (last q)': [],
    'Free Cash Flow': [],
    'Free Cash Flow Growth y/y': [],
    'Valuation': [],
    'P/FCF Ratio': [],
    'P/E Ratio': [],
    'P/S Ratio': [],
}

class TTM_Data(object):
    def __init__(self, stock_data_key, api_field_name):
        self.stock_data_key = stock_data_key
        self.api_field_name = api_field_name
        self.value = 0
        self.data_in_error = False

def execute(stock_tickers_input, api_key):
    # Since this can be run multiple times, make sure to clear array data
    for key in stock_data.keys():
        stock_data[key] = []

    stock_tickers = stock_tickers_input.split(',')

    # Uncomment for local dev
    # income_quarterly_reports = None
    # income_annual_reports = None
    # with open('static_json_data/apha_income.json') as json_file:
    #     income_statement = json.load(json_file)
    #     income_quarterly_reports = income_statement['quarterlyReports']
    #     income_annual_reports = income_statement['annualReports']
    #
    # balance_quarterly_reports = None
    # balance_annual_reports = None
    # with open('static_json_data/apha_balance.json') as json_file:
    #     balance_statement = json.load(json_file)
    #     balance_quarterly_reports = balance_statement['quarterlyReports']
    #     balance_annual_reports = balance_statement['annualReports']
    #
    # cashflow_quarterly_reports = None
    # cashflow_annual_reports = None
    # with open('static_json_data/apha_cashflow.json') as json_file:
    #     cashflow_statement = json.load(json_file)
    #     cashflow_quarterly_reports = cashflow_statement['quarterlyReports']
    #     cashflow_annual_reports = cashflow_statement['annualReports']
    #
    # overview_report = None
    # with open('static_json_data/apha_overview.json') as json_file:
    #     overview_report = json.load(json_file)

    stock_api_wrapper = stock_api.StockApiWrapper(api_key)
    heading_row = ["(#'s in millions OR %'s')"]
    # TODO: Actually implement calling all tickers in the stock api and get mapping of data
    count = 1
    for ticker in stock_tickers:
        print('picking up: {}'.format(ticker))
        heading_row.append(ticker)
        income_report = stock_api_wrapper.get_income_statement_json(ticker)
        income_quarterly_reports = income_report['quarterlyReports']
        income_annual_reports = income_report['annualReports']

        cashflow_report = stock_api_wrapper.get_cashflow_json(ticker)
        cashflow_quarterly_reports = cashflow_report['quarterlyReports']
        cashflow_annual_reports = cashflow_report['annualReports']

        balance_report = stock_api_wrapper.get_balance_json(ticker)
        balance_quarterly_reports = balance_report['quarterlyReports']
        balance_annual_reports = balance_report['annualReports']

        overview_report = stock_api_wrapper.get_company_overview_json(ticker)

        ttm_income_data_list = {
            'Revenue': TTM_Data('Revenue TTM', 'totalRevenue'),
            'Cost of Revenue': TTM_Data('Cost of Revenue TTM', 'costOfRevenue'),
            'Gross Profit': TTM_Data('Gross Profit TTM', 'grossProfit'),
            'Operating Income': TTM_Data('Operating Income TTM', 'operatingIncome'),
        }
        for qr in income_quarterly_reports[:4]:
            for ttm_data in ttm_income_data_list.values():
                value = excel_helpers.string_to_float(qr[ttm_data.api_field_name])
                if value == 'ERROR':
                    ttm_data.data_in_error = True
                    print('ERROR FOUND')
                    print(ticker)
                    print(ttm_data)
                else:
                    ttm_data.value = ttm_data.value + value

        for ttm_data in ttm_income_data_list.values():
            new_value = ttm_data.value / 1000000
            if ttm_data.data_in_error:
                new_value = '{}***ERROR'.format(new_value)
            stock_data[ttm_data.stock_data_key].append(new_value)

        # for some reason the api doesn't return this for us
        operating_expense = ttm_income_data_list['Gross Profit'].value - ttm_income_data_list['Operating Income'].value
        stock_data['Operating Expense TTM'].append(operating_expense / 1000000)

        gross_margin = ttm_income_data_list['Gross Profit'].value / ttm_income_data_list['Revenue'].value
        stock_data['Gross Margin TTM'].append(round(gross_margin, 3))

        operating_margin = ttm_income_data_list['Operating Income'].value / ttm_income_data_list['Revenue'].value
        stock_data['Operating Margin TTM'].append(round(operating_margin, 3))

        revenue_year_one = excel_helpers.string_to_float(income_annual_reports[0]['totalRevenue'])
        revenue_year_two = excel_helpers.string_to_float(income_annual_reports[1]['totalRevenue'])
        renevue_growth_yoy = (revenue_year_one - revenue_year_two) / revenue_year_two
        stock_data['Revenue Growth y/y'].append(round(renevue_growth_yoy, 3))

        most_recent_balance_q = balance_quarterly_reports[0]
        previous_balance_q = balance_quarterly_reports[1]
        stock_data['Total Cash'].append(excel_helpers.string_to_float(most_recent_balance_q['cash']) / 1000000)

        total_current_assets = excel_helpers.string_to_float(most_recent_balance_q['totalCurrentAssets'])
        stock_data['Current Assets'].append(total_current_assets / 1000000)

        current_liabilities = excel_helpers.string_to_float(most_recent_balance_q['totalCurrentLiabilities'])
        stock_data['Current Liabilities Total'].append(current_liabilities / 1000000)

        current_ratio = round(total_current_assets / current_liabilities, 3)
        stock_data['Current Ratio'].append(current_ratio)

        total_assets = excel_helpers.string_to_float(most_recent_balance_q['totalAssets'])
        stock_data['Total Assets'].append(total_assets / 1000000)

        total_liabilities = excel_helpers.string_to_float(most_recent_balance_q['totalLiabilities'])
        stock_data['Total Liabilities'].append(total_liabilities / 1000000)

        stock_holder_equity = total_assets - total_liabilities
        stock_data['Stock Holder Equity'].append(stock_holder_equity / 1000000)

        total_assets_previous_q = excel_helpers.string_to_float(previous_balance_q['totalAssets'])
        total_liabilites_previous_q = excel_helpers.string_to_float(previous_balance_q['totalLiabilities'])
        stock_holder_equity_previous_q = total_assets_previous_q - total_liabilites_previous_q
        stock_holder_growth_qoq = round((stock_holder_equity - stock_holder_equity_previous_q) / stock_holder_equity_previous_q, 3)
        stock_data['Stock Holder Growth q/q'].append(stock_holder_growth_qoq)

        goodwill = excel_helpers.string_to_float(most_recent_balance_q['goodwill'])
        if goodwill == 'ERROR':
            goodwill = 0
        stock_data['Goodwill'].append(goodwill / 1000000)
        intangible_assets = excel_helpers.string_to_float(most_recent_balance_q['intangibleAssets'])
        if intangible_assets == 'ERROR':
            intangible_assets = 0
        stock_data['Intangible Assets'].append(intangible_assets / 1000000)

        tangible_book_value = excel_helpers.string_to_float(most_recent_balance_q['netTangibleAssets'])
        stock_data['Tangible Book Value'].append(tangible_book_value / 1000000)
        stock_data['Tangible Assets'].append((tangible_book_value  + total_liabilities) / 1000000)

        # operatingCashflow - capitalExpenditures
        most_recent_cashflow_q = cashflow_quarterly_reports[0]
        operating_cashflow = excel_helpers.string_to_float(most_recent_cashflow_q['operatingCashflow'])
        capital_expenditures = excel_helpers.string_to_float(most_recent_cashflow_q['capitalExpenditures'])
        stock_data['Free Cash Flow (last q)'].append((operating_cashflow - capital_expenditures) / 1000000)

        most_recent_cashflow_y = cashflow_annual_reports[0]
        previous_cashflow_y = cashflow_annual_reports[1]
        operating_cashflow_recent = excel_helpers.string_to_float(most_recent_cashflow_y['operatingCashflow'])
        capital_expenditures_recent = excel_helpers.string_to_float(most_recent_cashflow_y['capitalExpenditures'])
        operating_cashflow_previous = excel_helpers.string_to_float(previous_cashflow_y['operatingCashflow'])
        capital_expenditures_previous = excel_helpers.string_to_float(previous_cashflow_y['capitalExpenditures'])
        free_cash_flow_recent = operating_cashflow_recent - capital_expenditures_recent
        free_cash_flow_previous = operating_cashflow_previous - capital_expenditures_previous
        stock_data['Free Cash Flow Growth y/y'].append(round((free_cash_flow_recent - free_cash_flow_previous) / free_cash_flow_previous, 3))

        operating_cashflow_ttm = 0
        capital_expenditure_ttm = 0
        for report in cashflow_quarterly_reports[:4]:
            operating_cashflow_ttm = operating_cashflow_ttm + excel_helpers.string_to_float(report['operatingCashflow'])
            capital_expenditure_ttm = capital_expenditure_ttm + excel_helpers.string_to_float(report['capitalExpenditures'])
        stock_data['Free Cash Flow'].append((operating_cashflow_ttm - capital_expenditure_ttm) / 1000000)

        market_cap = excel_helpers.string_to_float(overview_report['MarketCapitalization'])
        stock_data['Valuation'].append(market_cap / 1000000)
        stock_data['P/FCF Ratio'].append('TODO')
        stock_data['P/E Ratio'].append(excel_helpers.string_to_float(overview_report['PERatio']))
        stock_data['P/S Ratio'].append(excel_helpers.string_to_float(overview_report['PriceToSalesRatioTTM']))

        print('finished {} sleeping for 60'.format(ticker))
        if count != len(stock_tickers):
            time.sleep(60)  # alpha vantage limits us to 5 api calls per minute :( #DontWannaPay
        count = count + 1

    workbook = xl.Workbook()
    worksheet = workbook.create_sheet('Overview', 0)
    worksheet.append(heading_row)
    for key in stock_data.keys():
        row_data = [key] + stock_data[key]
        worksheet.append(row_data)

    columns_length = len(stock_tickers) + 1
    worksheet.add_table(xltable.Table(
        displayName='Overview',
        # If this isn't specified the program will crash in a cryptic way
        ref='A1:{}{}'.format(
            xlutils.get_column_letter(columns_length),
            len(stock_data.keys()) + 1,
        ),
        tableStyleInfo=xltable.TableStyleInfo(
            name='TableStyleMedium3',
            showFirstColumn=True,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
    ))

    for i in range(columns_length):
        worksheet.column_dimensions[xlutils.get_column_letter(i + 1)].width = 25

    workbook.save('pronkSheet.xlsx')
    os.startfile('pronkSheet.xlsx')
